import pywt
import numpy as np
from scipy import signal
from openpyxl import load_workbook


path = 'C:/Users/JulianD Colorado/Desktop/matlab-scripts/'
book_excel = load_workbook(path + 'Filtrados XYZ 1D (Ref ABCC).xlsx')
wb_sheets = book_excel.get_sheet_names()
print(wb_sheets[1])
sheet_ranges = book_excel[wb_sheets[1]]
print(sheet_ranges)
column_size = len(sheet_ranges['A'])
print(column_size)

vector_dateX = []
vector_X = []
vector_SigmaX = []
for x in range(2,(column_size+1)):   
    ## print(sheet_ranges.cell(row = x, column = 1).value)
    vector_dateX.append(sheet_ranges.cell(row = x, column = 1).value)
    vector_X.append(sheet_ranges.cell(row = x, column = 2).value)
    vector_SigmaX.append(sheet_ranges.cell(row = x, column = 3).value)

vector_dateX = np.array([vector_dateX])
vector_X = np.array([vector_X])
vector_SigmaX = np.array([vector_SigmaX])

coef,freqs=pywt.cwt()
